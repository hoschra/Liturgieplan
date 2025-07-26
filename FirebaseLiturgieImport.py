# -*- coding: utf-8 -*-
"""
Created on Sat Apr 17 07:13:21 2021
@author: Volker.Erben@gmail.com

remark close Excel fiel before start 
"""
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
from openpyxl import load_workbook, Workbook
import time
from os import path
from datetime import datetime
import re
from google.api_core.datetime_helpers import DatetimeWithNanoseconds
import pytz
import string
import random


result=[]

def main():
    filename=r"h:\Meine Ablage\Kirche-Liturgieplan\Liturgieplan.xlsx"  
    workbook=ExcelLoad(filename)
    FirebaseDelete() 
    ReadCurrentPlan(workbook, "Apr25")
    
def FirebaseDelete():    
    tempAppName = "Kasperd"+time.strftime("%H%M%S", time.localtime())
    cred = credentials.Certificate('pfarrhelfer-a4d48-56b56c899081.json')
    try:
        firebase_admin.initialize_app(cred)
    except:     
        firebase_admin.initialize_app(cred,name=tempAppName)
    db = firestore.client()
    #delete old one
    docs = db.collection(u'entry').stream()
    for doc in docs:
        db.collection(u'entry').document(doc.id).delete()
    
def id_generator(size=6, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


def FirebaseImport():
    global result
    tempAppName = "Kasper"+time.strftime("%H%M%S", time.localtime())+id_generator()
    # Use the credentials from personal firebase
    cred = credentials.Certificate('pfarrhelfer-a4d48-56b56c899081.json')
    #unclear behavior will first time only work with this line
    try:
        firebase_admin.initialize_app(cred)
    except:     
        firebase_admin.initialize_app(cred,name=tempAppName)
    db = firestore.client()
    #delete old one
    docs = db.collection(u'entry').stream()
    local_time = pytz.timezone("Europe/Berlin")
    all=False
    for doc in docs:
        #print(f'{doc.id} => {doc.to_dict()}')
        if all:
            db.collection(u'entry').document(doc.id).delete()    
        else:            
            dict=doc.to_dict()
            #remove wrongly created
            if type(dict['date']) is str:
                #print(dict['date'])
                db.collection(u'entry').document(doc.id).delete()
            else:
                now = DatetimeWithNanoseconds.now(local_time)
                #print(temp)
                if dict['date'] < now:
                    #print(dict['date']+" deleted")
                    db.collection(u'entry').document(doc.id).delete()    
    
    #set new
    for entry in result:
        print (entry)
        doc_ref = db.collection(u'entry').document()
        doc_ref.set(entry)
        #break;
    
def ExcelLoad(excelfile):
    if path.exists(excelfile):
        wb = load_workbook(filename=excelfile)
    else:
        wb = Workbook()    
    return wb

def SelectSheet(wb,wbname):
    try:    
        ws = wb[wbname]
    except:
        print("page not exist {}".format(wbname))
        ws=-1
    return ws


def ReadCurrentPlan(wb,sheetname):
    global result
    ws=SelectSheet(wb,sheetname)
    if ws != -1:
        print("sheet found")
        #print(tuple(ws.rows))
        first=True
        for col in ws.iter_rows(min_row=40, max_col=4, max_row=57):
        #for col in ws.iter_rows(min_row=58, max_col=4, max_row=120):    
            entry={}
            if first:
                first=False
                #TODO find year
                #print(col[0].value)
            else:   
                print(type(col[0].value))
                if col[0].value != None:
                    print(f"a {col[0].value} \nGDH1: {col[1].value}")
                    entry['merge']=0
                    entry['type']='GDH'
                    entry['date']=ConvertDate(col[0].value,2024)
                    entry['sup']=col[1].value
                    entry['color']='yellow'
                    result.append(entry)
                    if type(col[2].value) is str:
                        if len(col[2].value):
                            print (f"GDH2: {col[2].value}" )
                            entry={}
                            entry['merge']=0
                            entry['type']='GDH'
                            entry['date']=ConvertDate(col[0].value,2024)
                            entry['sup']=col[2].value
                            entry['color']='yellow'
                            result.append(entry)
                    if type(col[3].value) is str:
                        if len(col[3].value):
                            print (f"kommunionhelfer: {col[3].value}" )
                            entry={}
                            entry['merge']=0
                            entry['type']='GDH'
                            entry['date']=ConvertDate(col[0].value,2024)
                            entry['sup']=col[3].value
                            entry['color']='white'
                            result.append(entry)
            if len(result)>10:
                FirebaseImport();
                result=[]
    FirebaseImport();     
    return 

def ConvertDate(datestring,year):
    
    monthdict =  { "Jan":"1", "Feb":"2", "Mar":"3", "Mrz":"3" ,"Apr":"4","April":"4" , "Mai":"05", "Jun":"6","Jul":"7", "Aug":"8", "Sep":"9","Okt":"10", "Nov":"11", "Dez":"12"}
    local_time  = pytz.timezone("Europe/Berlin")
    #print (datestring)
    pattern1=re.search(r'(\d\d).(\w+)\s(\d+):(\d\d)',datestring) 
    if pattern1:
        #print ("{}.{}.{} {}:{}".format(pattern1.group(1),monthdict[pattern1.group(2)],year,pattern1.group(3),pattern1.group(4)))
        date_time_str=pattern1.group(1)+"/"+monthdict[pattern1.group(2)]+"/"+str(year)+" "+pattern1.group(3)+":"+pattern1.group(4)+":00"
        naive_datetime = datetime. strptime(date_time_str, '%d/%m/%Y %H:%M:%S')
        local_datetime = local_time.localize(naive_datetime, is_dst=None)
        utc_datetime = local_datetime.astimezone(pytz.utc)
        #0001-01-01T00:00:00Z to 9999-12-31T23:59:59.999999999Z
        #retval=utc_datetime.isoformat("T")
        #retval=retval[:-6]+".000000000Z"
        
    else:
        print ("Can't match date: {}".format(datestring))
    return utc_datetime
        
if __name__ == '__main__':
   main()        
